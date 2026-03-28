"""
parse_student_roster.py
-----------------------
Reads 3 student roster Excel files (one per grade) and produces:
  - students.json : one entry per student with NIS, name, gender, class

File structure (all 3 files identical layout):
  Sheet1 contains 12 class blocks stacked vertically.
  Each block: 6 header rows (school name, title, etc) + data rows.
  Header detected by col A = 'NO'.
  Student rows: col A = number, col B = NIS, col C = NISN, col D = name, col E = L/P.
  No class name in data — classes are in alphabetical order within grade.

Class ordering:
  Grade 10: X - A, X - B, ..., X - L
  Grade 11: XI - A, XI - B, ..., XI - L
  Grade 12: XII IPA - 1..8, XII IPS - 1..4
"""

import json
from openpyxl import load_workbook

# ── Configuration ───────────────────────────────────────────────────────────

ROSTER_FILES = [
    {
        "grade": 10,
        "file": "Daftar_Hadir_Siswa_kelas_X_2025-2026-NEW.xlsx",
        "classes": [
            "X - A", "X - B", "X - C", "X - D", "X - E", "X - F",
            "X - G", "X - H", "X - I", "X - J", "X - K", "X - L",
        ],
    },
    {
        "grade": 11,
        "file": "Daftar_Hadir_Siswa_kelas_XI_2025-2026_-NEW.xlsx",
        "classes": [
            "XI - A", "XI - B", "XI - C", "XI - D", "XI - E", "XI - F",
            "XI - G", "XI - H", "XI - I", "XI - J", "XI - K", "XI - L",
        ],
    },
    {
        "grade": 12,
        "file": "Daftar_Hadir_Siswa_kelas_XII_2025-2026-NEW.xlsx",
        "classes": [
            "XII IPA - 1", "XII IPA - 2", "XII IPA - 3", "XII IPA - 4",
            "XII IPA - 5", "XII IPA - 6", "XII IPA - 7", "XII IPA - 8",
            "XII IPS - 1", "XII IPS - 2", "XII IPS - 3", "XII IPS - 4",
        ],
    },
]

# Update this to your actual file directory
INPUT_DIR = "."
OUTPUT_FILE = "students.json"


# ── Parser ──────────────────────────────────────────────────────────────────

def parse_single_file(filepath, class_names):
    """
    Parse one roster Excel file.
    Returns list of (class_name, students_list) tuples.
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find all header rows (col A == 'NO')
    header_indices = []
    for i, row in enumerate(rows):
        if row[0] and str(row[0]).strip() == "NO":
            header_indices.append(i)

    if len(header_indices) != len(class_names):
        raise ValueError(
            f"Expected {len(class_names)} class blocks in {filepath}, "
            f"found {len(header_indices)} header rows"
        )

    results = []
    for block_idx, header_row in enumerate(header_indices):
        # Students start 3 rows after header (header + sub-header + blank)
        start = header_row + 3

        # Students end before next header block (minus ~6 rows of header)
        # or at end of file
        if block_idx + 1 < len(header_indices):
            end = header_indices[block_idx + 1] - 3  # safe margin before next header
        else:
            end = len(rows)

        students = []
        for i in range(start, end):
            row = rows[i]
            no_val = row[0]

            # Skip non-student rows
            if no_val is None:
                continue
            try:
                int(str(no_val))
            except (ValueError, TypeError):
                continue

            nis = str(row[1]).strip() if row[1] else None
            nisn = str(row[2]).strip() if row[2] else None
            name = str(row[3]).strip() if row[3] else None
            gender = str(row[4]).strip() if row[4] else None

            if nis and name:
                students.append({
                    "nis": nis,
                    "nisn": nisn,
                    "name": name,
                    "gender": gender,  # L or P
                })

        results.append((class_names[block_idx], students))

    return results


def main():
    all_students = []
    total = 0

    for config in ROSTER_FILES:
        filepath = f"{INPUT_DIR}/{config['file']}"
        grade = config["grade"]
        class_names = config["classes"]

        print(f"Parsing grade {grade}: {config['file']}")
        class_data = parse_single_file(filepath, class_names)

        for class_name, students in class_data:
            print(f"  {class_name}: {len(students)} students")
            for s in students:
                all_students.append({
                    "nis": s["nis"],
                    "nisn": s["nisn"],
                    "name": s["name"],
                    "gender": s["gender"],
                    "class": class_name,
                    "grade": grade,
                })
            total += len(students)

        print()

    # Write output
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(all_students, f, ensure_ascii=False, indent=2)

    print(f"Total: {total} students written to {OUTPUT_FILE}")

    # Sanity check: duplicate NIS?
    nis_list = [s["nis"] for s in all_students]
    nis_set = set(nis_list)
    if len(nis_list) != len(nis_set):
        dupes = [n for n in nis_set if nis_list.count(n) > 1]
        print(f"⚠ WARNING: {len(dupes)} duplicate NIS found: {dupes[:5]}")
    else:
        print("✓ All NIS values unique")


if __name__ == "__main__":
    main()
